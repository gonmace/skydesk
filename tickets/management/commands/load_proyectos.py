"""Carga/actualiza Proyectos desde un Excel (hoja "Base de datos ").

    python manage.py load_proyectos [--file proyectos_bd.xlsx] [--dry-run]

Idempotente: hace update_or_create por `code`, así que se puede volver a
correr para sincronizar cambios del Excel sin duplicar proyectos.
"""
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError

from tickets.models import Project

SHEET_NAME = 'Base de datos '

STATUS_MAP = {
    'EN CURSO': Project.Status.ACTIVE,
    'SIN INICIAR': Project.Status.ACTIVE,
    'STAND BY': Project.Status.PAUSED,
    'COMPLETO': Project.Status.CLOSED,
    'DE BAJA': Project.Status.CLOSED,
}

REGIONAL_CITY = {
    'CBB': 'Cochabamba',
    'LPZ': 'La Paz',
    'SCZ': 'Santa Cruz',
    'TJA': 'Tarija',
}


class Command(BaseCommand):
    help = 'Carga proyectos desde proyectos_bd.xlsx (hoja "Base de datos ") a Project.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', default=str(Path(settings.BASE_DIR) / 'proyectos_bd.xlsx'),
            help='Ruta al .xlsx (default: proyectos_bd.xlsx en la raíz del proyecto).',
        )
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Muestra qué se haría sin escribir en la base de datos.',
        )

    def handle(self, *args, **options):
        path = Path(options['file'])
        if not path.exists():
            raise CommandError(f'No se encontró el archivo: {path}')

        wb = openpyxl.load_workbook(path, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            raise CommandError(
                f'No se encontró la hoja "{SHEET_NAME}". Hojas disponibles: {wb.sheetnames}'
            )
        ws = wb[SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))[1:]  # salta encabezado

        created = updated = skipped = 0
        for i, row in enumerate(rows, start=2):
            if not any(c is not None for c in row):
                continue

            code = (row[2] or '').strip().upper() if row[2] else ''
            name = (row[7] or '').strip() if row[7] else ''
            regional = (row[4] or '').strip().upper() if row[4] else ''
            estatus = (row[20] or '').strip().upper() if row[20] else ''

            if not code or not name:
                self.stderr.write(self.style.WARNING(f'Fila {i}: sin código o nombre, se omite.'))
                skipped += 1
                continue

            status = STATUS_MAP.get(estatus, Project.Status.ACTIVE)
            city = REGIONAL_CITY.get(regional, regional)

            if options['dry_run']:
                exists = Project.objects.filter(code=code).exists()
                self.stdout.write(f'{"UPDATE" if exists else "CREATE"} {code} · {name} ({city}, {status})')
                continue

            _, was_created = Project.objects.update_or_create(
                code=code,
                defaults={'name': name, 'city': city, 'status': status},
            )
            if was_created:
                created += 1
            else:
                updated += 1

        if options['dry_run']:
            self.stdout.write(self.style.SUCCESS(f'Dry-run: {len(rows) - skipped} proyectos procesados, {skipped} omitidos.'))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'{created} proyectos creados, {updated} actualizados, {skipped} omitidos.'
            ))
